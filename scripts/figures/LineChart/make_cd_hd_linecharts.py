from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib.ticker import FormatStrFormatter
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
OUTPUT_DIR = REPO_ROOT / "results" / "figures"
ROUND_AUDIT_CSV = REPO_ROOT / "results" / "summaries" / "geometry" / "dimension1_round_compile_audit_20260607.csv"
REFERENCE_XLSX = REPO_ROOT / "docs" / "Metrics_References.xlsx"

OUTPUT_PNG = OUTPUT_DIR / "cd_hd_iteration_linecharts.png"
OUTPUT_SVG = OUTPUT_DIR / "cd_hd_iteration_linecharts.svg"
OUTPUT_VALUES_CSV = OUTPUT_DIR / "cd_hd_iteration_values.csv"
OUTPUT_REFERENCE_CSV = OUTPUT_DIR / "cd_hd_reference_check.csv"
OUTPUT_EFFECTIVE_CSV = OUTPUT_DIR / "iterative_effective_materialization_check.csv"


ROUND_ORDER = ["One-shot", "Iteration1", "Iteration2", "Iteration3"]
ROUND_LABELS = ["One-shot", "Iter. 1", "Iter. 2", "Iter. 3"]

NON_ITERATIVE_METHODS = ["DeepSeek", "Codex-MCP", "Text-to-CadQuery", "Text2CAD"]
ITERATIVE_METHODS = ["Claude", "ChatGPT", "Qwen", "Gemini"]
METHOD_ORDER = NON_ITERATIVE_METHODS + ITERATIVE_METHODS

DISPLAY_NAME = {
    "DeepSeek": "DeepSeek-v4 Flash",
    "Codex-MCP": "Codex-MCP",
    "Text-to-CadQuery": "Text-to-CadQuery",
    "Text2CAD": "Text2CAD",
    "Claude": "Claude",
    "ChatGPT": "ChatGPT",
    "Qwen": "Qwen",
    "Gemini": "Gemini",
}

LINE_LABEL = {
    "DeepSeek": "DeepSeek-v4",
    "Codex-MCP": "Codex-MCP",
    "Text-to-CadQuery": "Text-to-CadQuery",
    "Text2CAD": "Text2CAD",
    "Claude": "Claude",
    "ChatGPT": "ChatGPT",
    "Qwen": "Qwen",
    "Gemini": "Gemini",
}

LINE_COLORS = {
    "DeepSeek": "#0B6E4F",
    "Codex-MCP": "#2A9D8F",
    "Text-to-CadQuery": "#8A5A44",
    "Text2CAD": "#C06C3E",
    "Claude": "#315C9A",
    "ChatGPT": "#4F83CC",
    "Qwen": "#7A5195",
    "Gemini": "#B05AA0",
}

COLORS = {
    "dark": "#0B4F2A",
    "grid": "#D9E4DC",
    "muted": "#6B7280",
    "text": "#111827",
    "panel": "#F8FBF8",
    "border": "#A7BDAE",
}


def per_round_metrics(csv_path: Path) -> tuple[float, float, int]:
    cds: list[float] = []
    hds: list[float] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if str(row.get("compile_success", "")).strip() != "1":
                continue
            cd_raw = row.get("chamfer_distance", "")
            hd_raw = row.get("hausdorff_distance", "")
            if cd_raw in (None, "") or hd_raw in (None, ""):
                continue
            cds.append(float(cd_raw))
            hds.append(float(hd_raw))
    if not cds or not hds:
        raise ValueError(f"No valid CD/HD values found in {csv_path}")
    return sum(cds) / len(cds), sum(hds) / len(hds), len(cds)


def repo_path(value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else REPO_ROOT / path


def display_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return str(path)


def read_round_values() -> list[dict[str, object]]:
    with ROUND_AUDIT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        audit_rows = list(csv.DictReader(f))

    rows: list[dict[str, object]] = []
    for audit in audit_rows:
        method = audit["method"]
        round_name = audit["round"]
        if method not in METHOD_ORDER or round_name not in ROUND_ORDER:
            continue
        per_sample_csv = repo_path(audit["per_sample_csv"])
        mean_cd, mean_hd, valid_outputs = per_round_metrics(per_sample_csv)
        rows.append(
            {
                "method": method,
                "display_method": DISPLAY_NAME[method],
                "cohort": audit["cohort"],
                "round": round_name,
                "round_index": ROUND_ORDER.index(round_name),
                "mean_cd": mean_cd,
                "mean_hd": mean_hd,
                "valid_outputs": valid_outputs,
                "success_from_audit": int(audit["success"]),
                "attempts": int(audit["attempts"]),
                "per_sample_csv": display_path(per_sample_csv),
            }
        )
    return rows


def write_round_values_csv(rows: list[dict[str, object]]) -> None:
    fieldnames = [
        "method",
        "display_method",
        "cohort",
        "round",
        "round_index",
        "mean_cd",
        "mean_hd",
        "valid_outputs",
        "success_from_audit",
        "attempts",
        "per_sample_csv",
    ]
    with OUTPUT_VALUES_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for method in METHOD_ORDER:
            method_rows = [r for r in rows if r["method"] == method]
            method_rows.sort(key=lambda r: int(r["round_index"]))
            for row in method_rows:
                writer.writerow(row)


def read_reference_geometry() -> dict[str, dict[str, object]]:
    wb = load_workbook(REFERENCE_XLSX, data_only=True, read_only=True)
    ws = wb["Geometry"]
    headers = [cell.value for cell in ws[1]]
    reference: dict[str, dict[str, object]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        method = row.get("method")
        if method in METHOD_ORDER:
            reference[str(method)] = row
    return reference


def write_reference_check(rows: list[dict[str, object]]) -> bool:
    if not REFERENCE_XLSX.exists():
        print(f"Reference workbook not included; skipping {OUTPUT_REFERENCE_CSV}")
        return False
    reference = read_reference_geometry()
    by_method: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        by_method.setdefault(str(row["method"]), []).append(row)

    fieldnames = [
        "method",
        "cohort",
        "xlsx_geometry_mean_cd",
        "xlsx_geometry_mean_hd",
        "min_round_mean_cd",
        "min_round_mean_hd",
        "iteration3_mean_cd",
        "iteration3_mean_hd",
        "reference_note",
    ]
    with OUTPUT_REFERENCE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for method in METHOD_ORDER:
            method_rows = by_method[method]
            method_rows.sort(key=lambda r: int(r["round_index"]))
            ref = reference.get(method, {})
            writer.writerow(
                {
                    "method": method,
                    "cohort": method_rows[0]["cohort"],
                    "xlsx_geometry_mean_cd": ref.get("mean_cd", ""),
                    "xlsx_geometry_mean_hd": ref.get("mean_hd", ""),
                    "min_round_mean_cd": min(float(r["mean_cd"]) for r in method_rows),
                    "min_round_mean_hd": min(float(r["mean_hd"]) for r in method_rows),
                    "iteration3_mean_cd": next(float(r["mean_cd"]) for r in method_rows if r["round"] == "Iteration3"),
                    "iteration3_mean_hd": next(float(r["mean_hd"]) for r in method_rows if r["round"] == "Iteration3"),
                    "reference_note": ref.get("geometry_source_summary", ""),
                }
            )
    return True


def md5(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def prediction_dir(method: str, round_name: str) -> Path:
    model_key = {
        "ChatGPT": "chatgpt",
        "Claude": "claude",
        "Gemini": "gemini",
        "Qwen": "qwen",
    }[method]
    return (
        REPO_ROOT
        / "excluded_full_output_tree"
        / "Results"
        / "Benchmark"
        / "Closed-loop Workflow"
        / round_name
        / method
        / "benchmark_input"
        / "predictions"
        / model_key
    )


def write_effective_materialization_check() -> bool:
    if not prediction_dir("ChatGPT", "One-shot").exists():
        print(f"Full STL output tree not included; skipping {OUTPUT_EFFECTIVE_CSV}")
        return False
    fieldnames = ["method", "comparison", "common_one_shot_outputs", "identical_to_one_shot_outputs"]
    with OUTPUT_EFFECTIVE_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for method in ITERATIVE_METHODS:
            one_shot_dir = prediction_dir(method, "One-shot")
            one_shot_hashes = {p.name: md5(p) for p in one_shot_dir.glob("*.stl")}
            for round_name in ["Iteration1", "Iteration2", "Iteration3"]:
                later_dir = prediction_dir(method, round_name)
                common = 0
                same = 0
                for name, file_hash in one_shot_hashes.items():
                    later = later_dir / name
                    if not later.exists():
                        continue
                    common += 1
                    if md5(later) == file_hash:
                        same += 1
                writer.writerow(
                    {
                        "method": method,
                        "comparison": f"One-shot vs {round_name}",
                        "common_one_shot_outputs": common,
                        "identical_to_one_shot_outputs": same,
                    }
                )
    return True


def rows_for_method(rows: list[dict[str, object]], method: str) -> list[dict[str, object]]:
    selected = [r for r in rows if r["method"] == method]
    selected.sort(key=lambda r: int(r["round_index"]))
    if len(selected) != 4:
        raise ValueError(f"Expected four rounds for {method}, found {len(selected)}")
    return selected


def value_range(rows: list[dict[str, object]], metric: str) -> tuple[float, float]:
    values = [float(r[metric]) for r in rows]
    span = max(values) - min(values)
    pad = max(span * 0.12, 0.004 if metric == "mean_cd" else 0.012)
    return max(0.0, min(values) - pad), max(values) + pad


def plot_group(ax, rows: list[dict[str, object]], methods: list[str], metric: str, title: str, y_label: str) -> None:
    x = list(range(4))
    for method in methods:
        method_rows = rows_for_method(rows, method)
        values = [float(r[metric]) for r in method_rows]
        ax.plot(
            x,
            values,
            color=LINE_COLORS[method],
            linewidth=2.4,
            marker="o",
            markersize=5.2,
            label=DISPLAY_NAME[method],
        )

    ax.set_facecolor(COLORS["panel"])
    ax.set_xticks(x, ROUND_LABELS)
    ax.set_ylabel(y_label, fontsize=28, labelpad=14)
    ax.yaxis.set_major_formatter(FormatStrFormatter("%.3f"))
    ax.grid(axis="y", color=COLORS["grid"], linewidth=0.9)
    ax.grid(axis="x", color="#EEF3EF", linewidth=0.5)
    ax.tick_params(axis="x", labelsize=26, pad=9)
    ax.tick_params(axis="y", labelsize=26, pad=9)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color(COLORS["dark"])
    ax.spines["bottom"].set_color(COLORS["dark"])


def add_line_end_labels(ax, rows: list[dict[str, object]], methods: list[str], metric: str) -> None:
    final_points: list[tuple[str, float]] = []
    for method in methods:
        method_rows = rows_for_method(rows, method)
        final_points.append((method, float(method_rows[-1][metric])))

    ymin, ymax = ax.get_ylim()
    span = ymax - ymin
    min_gap = span * 0.075
    adjusted: list[tuple[str, float, float]] = []
    last_y = ymin - span
    for method, y in sorted(final_points, key=lambda item: item[1]):
        label_y = max(y, last_y + min_gap)
        adjusted.append((method, y, label_y))
        last_y = label_y
    overflow = adjusted[-1][2] - (ymax - min_gap * 0.2)
    if overflow > 0:
        adjusted = [(method, y, label_y - overflow) for method, y, label_y in adjusted]

    for method, y, label_y in adjusted:
        color = LINE_COLORS[method]
        ax.annotate(
            LINE_LABEL[method],
            xy=(3, y),
            xytext=(3.16, label_y),
            textcoords="data",
            color=color,
            fontsize=20,
            fontweight="bold",
            va="center",
            ha="left",
            arrowprops={
                "arrowstyle": "-",
                "color": color,
                "lw": 0.9,
                "alpha": 0.75,
                "shrinkA": 0,
                "shrinkB": 4,
            },
        )


def plot_linecharts(rows: list[dict[str, object]]) -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "axes.edgecolor": COLORS["dark"],
            "axes.labelcolor": COLORS["text"],
            "xtick.color": COLORS["text"],
            "ytick.color": COLORS["text"],
            "svg.fonttype": "none",
        }
    )

    fig, axes = plt.subplots(2, 2, figsize=(18.4, 11.3), sharex=True)
    fig.patch.set_facecolor("white")

    cd_min, cd_max = value_range(rows, "mean_cd")
    hd_min, hd_max = value_range(rows, "mean_hd")

    plot_group(axes[0, 0], rows, NON_ITERATIVE_METHODS, "mean_cd", "", "Mean CD")
    plot_group(axes[0, 1], rows, ITERATIVE_METHODS, "mean_cd", "", "Mean CD")
    plot_group(axes[1, 0], rows, NON_ITERATIVE_METHODS, "mean_hd", "", "Mean HD")
    plot_group(axes[1, 1], rows, ITERATIVE_METHODS, "mean_hd", "", "Mean HD")

    for ax in axes[0, :]:
        ax.set_ylim(cd_min, cd_max)
    for ax in axes[1, :]:
        ax.set_ylim(hd_min, hd_max)
    for ax in axes.ravel():
        ax.set_xlim(-0.15, 4.55)
    for ax in axes[:, 1]:
        ax.set_ylabel("")
        ax.tick_params(axis="y", left=False, labelleft=False)
        ax.spines["left"].set_visible(False)

    add_line_end_labels(axes[0, 0], rows, NON_ITERATIVE_METHODS, "mean_cd")
    add_line_end_labels(axes[0, 1], rows, ITERATIVE_METHODS, "mean_cd")
    add_line_end_labels(axes[1, 0], rows, NON_ITERATIVE_METHODS, "mean_hd")
    add_line_end_labels(axes[1, 1], rows, ITERATIVE_METHODS, "mean_hd")

    fig.text(0.29, 0.965, "Open-loop Workflow", ha="center", va="top", fontsize=30, fontweight="bold", color=COLORS["text"])
    fig.text(0.735, 0.965, "Closed-loop Workflow", ha="center", va="top", fontsize=30, fontweight="bold", color=COLORS["text"])

    fig.subplots_adjust(left=0.085, right=0.94, top=0.90, bottom=0.095, hspace=0.36, wspace=0.18)
    fig.savefig(OUTPUT_PNG, dpi=300, bbox_inches="tight")
    fig.savefig(OUTPUT_SVG, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_round_values()
    write_round_values_csv(rows)
    wrote_reference = write_reference_check(rows)
    wrote_effective = write_effective_materialization_check()
    plot_linecharts(rows)
    print(f"PNG written to: {OUTPUT_PNG}")
    print(f"SVG written to: {OUTPUT_SVG}")
    print(f"Round values written to: {OUTPUT_VALUES_CSV}")
    if wrote_reference:
        print(f"Reference check written to: {OUTPUT_REFERENCE_CSV}")
    if wrote_effective:
        print(f"Effective materialization check written to: {OUTPUT_EFFECTIVE_CSV}")


if __name__ == "__main__":
    main()
